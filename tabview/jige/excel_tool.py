import configparser
import os
import sys
import time
import re
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtWidgets import QLabel, QSpinBox, QHBoxLayout, QPushButton, QComboBox, QMessageBox, QVBoxLayout, QTextEdit
from openpyxl import load_workbook, Workbook

from control.myControl import myLineEdit
from .excel_tool_helper import copy_xls
from ..base import view as baseView
from ..thread import uiThread_mode2, uiThread
import asyncio



class view(baseView):
    def __init__(self,*args,**kwargs):
        super(view, self).__init__(*args,**kwargs)
        self.__pwd = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])),"jige")
        self.quit = False
        self.__init_ui()
        # self.timer = QTimer(self)
        # self.timer.timeout.connect(self.__second_task)
        # self.timer.start()
    
    # def __refresh_text_edit_sgl(self,args):
    #     path = args[0]
    #     try:
    #         # isopen = function.is_open(path)
    #         # if not isopen:
    #         with open(path,"r", encoding="utf-8") as rf:
    #             content = rf.read()
    #             if self.textEdit.toPlainText() != content:
    #                 self.textEdit.setText(content)
    #     except:
    #         pass

    # def __second_task(self):
    #     if len(self.combox.currentText()):
    #         self.__sencond_thread = uiThread((os.path.join(self.__pwd, self.combox.currentText()),))
    #         self.__sencond_thread.task_emit.connect(self.__refresh_text_edit_sgl)
    #         self.__sencond_thread.start()

    def __textEdit_textchange_sgl(self):
        asyncio.run(self.writeText2ini())
        print(self.textEdit.toPlainText())

    async def writeText2ini(self):
        with open(os.path.join(self.__pwd, self.combox.currentText()), "w", encoding="utf-8") as wf:
            wf.write(self.textEdit.toPlainText())

    def _box_index_change_sgl(self,index):
        self.__box_thread = uiThread((os.path.join(self.__pwd, self.combox.currentText()),))
        self.__box_thread.task_emit.connect(self.__refresh_text_edit_sgl)
        self.__box_thread.start()

    def __refresh_text_edit_sgl(self, args):
        path = args[0]
        with open(path,"r", encoding="utf-8") as rf:
            self.textEdit.setText(rf.read())

    def __init_ui(self):
        self.textEdit = QTextEdit()
        self.textEdit.setStyleSheet("""
            font-size: 14px;
            font-weight:bold;
        """)
        pt = QPalette()
        pt.setBrush(QPalette.Text, Qt.white)
        pt.setBrush(QPalette.Base, QColor("#3c3f41"))
        pt.setBrush(QPalette.Highlight, QColor("#d7686d"))
        self.textEdit.setPalette(pt)
        self.textEdit.textChanged.connect(self.__textEdit_textchange_sgl)
        label = QLabel("源excel      ")
        self.source_path_edit = myLineEdit()
        self.source_path_edit.setPlaceholderText(u"将源excel拖拽至此处")
        self.source_spinBox = QSpinBox()
        layout0 = QHBoxLayout()
        layout0.addWidget(label)
        layout0.addWidget(self.source_path_edit)
        layout0.addWidget(self.source_spinBox)
        label1 = QLabel("模板excel    ")
        self.mode_path_edit = myLineEdit()
        self.mode_path_edit.setPlaceholderText(u"将模板excel拖拽至此处")
        self.mode_spinBox = QSpinBox()
        layout1 = QHBoxLayout()
        layout1.addWidget(label1)
        layout1.addWidget(self.mode_path_edit)
        layout1.addWidget(self.mode_spinBox)
        label2 = QLabel("导出excel路径")
        self.export_path_edit = myLineEdit()
        self.export_path_edit.setPlaceholderText(u"将导出excel的路径拖拽至此处")
        layout2 = QHBoxLayout()
        layout2.addWidget(label2)
        layout2.addWidget(self.export_path_edit)
        layout3 = QHBoxLayout()
        self.excute_btn = QPushButton("导出")
        self.excute_btn.clicked.connect(self.__export_excel_sgl)
        self.quit_btn = QPushButton("中止")
        self.quit_btn.clicked.connect(self.__quit_export_sgl)
        layout3.addStretch(1)
        try:
            self.combox = QComboBox()
            self.combox.currentIndexChanged.connect(self._box_index_change_sgl)
            files = os.listdir(self.__pwd)
            for f in files:
                if f.endswith("ini"):
                    self.combox.addItem(f)
            self.combox.setCurrentIndex(0)
            with open(os.path.join(self.__pwd, self.combox.currentText()),"r",encoding="utf-8") as rf:
                content = rf.read()
                self.textEdit.setText(content)
            layout3.addWidget(self.combox)
        except Exception as e:
            QMessageBox.warning(self, "警告", f"{str(e)}")

        layout3.addWidget(self.excute_btn)
        layout3.addWidget(self.quit_btn)
        layout3.addStretch(1)

        self.log = QLabel("." * 100)
        layout4 = QHBoxLayout()
        layout4.addWidget(self.log)

        layout = QVBoxLayout()
        layout.addWidget(self.textEdit)
        layout.addLayout(layout0)
        layout.addLayout(layout1)
        layout.addLayout(layout2)
        layout.addLayout(layout3)
        layout.addLayout(layout4)
        layout.addStretch(0)
        self.setLayout(layout)

    def __quit_export_sgl(self):
        if hasattr(self, "myThread"):
            # self.myThread.quit()
            self.quit = True
        self.excute_btn.setEnabled(True)

    def __export_excel_sgl(self):
        self.quit = False
        self.myThread = uiThread_mode2(self.__generate_excel)
        self.myThread.start()

    def __generate_excel(self):
        boxtext = self.combox.currentText()
        if not len(boxtext):
            return QMessageBox.warning(self, "警告", f"ini不能为空。")
        source_path = self.source_path_edit.text()
        mode_path = self.mode_path_edit.text()
        export_path = self.export_path_edit.text()
        if not os.path.exists(source_path):
            return QMessageBox.warning(self, "警告", f"不存在源excel路径:{source_path}")
        if not os.path.exists(mode_path):
            return QMessageBox.warning(self, "警告", f"不存在模板excel路径:{mode_path}")
        if not os.path.exists(export_path):
            return QMessageBox.warning(self, "警告", f"不存在导出excel路径:{export_path}")
        bn1 = os.path.basename(source_path)
        bn2 = os.path.basename(mode_path)
        if not os.path.isfile(source_path) or not (bn1.endswith("xls") or bn1.endswith("xlsx")):
            return QMessageBox.warning(self, "警告", f"源excel不是一个文件或者excel:{source_path}")
        if not os.path.isfile(mode_path) or not (bn2.endswith("xls") or bn2.endswith("xlsx")):
            return QMessageBox.warning(self, "警告", f"模板excel不是一个文件或者excel:{source_path}")
        if not os.path.isdir(export_path):
            return QMessageBox.warning(self, "警告", f"导出excel路径不是一个目录:{source_path}")

        try:
            ini_config = configparser.ConfigParser()
            ini_config.read(os.path.join(self.__pwd, self.combox.currentText()))
            sheet_name_col = ini_config["config"]["sheet_name_col"]
            match = ini_config["match"]
            self.setEnabled(False)
            try:
                self.generate_excel(source_path, mode_path, export_path, sheet_name_col, dict(match),
                                    int(self.source_spinBox.text()),
                                    int(self.mode_spinBox.text()))
            except Exception as e:
                QMessageBox.critical(self, "excel转换错误", f"{str(e)}")
            self.setEnabled(True)
        except Exception as e:
            self.setEnabled(True)
            return QMessageBox.critical(self, "ini配置错误", f"{str(e)}")

    def generate_excel(self, source_path, mode_path, export_path, sheet_name_col, match, sIndex, mIndex):
        source_wb = load_workbook(source_path)
        start_row = int(sheet_name_col[-1])
        sn_mask = sheet_name_col[0:-1]
        source_sheet = source_wb.worksheets[sIndex]
        target_xlsx = Workbook()
        tagert_path = os.path.join(export_path, f"wrokbook{int(time.time())}.xlsx")
        target_xlsx.save(tagert_path)
        target_xlsx.close()
        src_ws = load_workbook(mode_path).worksheets[mIndex]
        wm = list(src_ws.merged_cells)  # 开始处理合并单元格
        rows = source_sheet[sn_mask]
        end_row = len(rows)
        for i, r in enumerate(rows):
            if i < start_row - 1:
                continue
            if r.value == None:
                end_row = i + 1
                break

        for i in range(start_row, end_row):
            if not self.quit:
                try:
                    sn = f"{sn_mask}{i}"
                    sheetname = source_sheet[sn].value
                    copy_xls(mode_path, tagert_path, sheetname, mIndex)
                    target_xlsx = load_workbook(tagert_path)
                    target_sheet = target_xlsx[sheetname]
                    self.log.setText(f"已完成{i + 1 - start_row}/{end_row - start_row},正在生成{sheetname}")
                    for k, v in match.items():
                        value = source_sheet[f"{k[0:-1]}{i}"].value
                        vsp = re.findall("[\d\w]+",v)
                        for vs in vsp:
                            target_sheet[vs] = value
                    for i in range(0, len(wm)):
                        cell2 = str(wm[i]).replace('(<MergedCellRange ', '').replace('>,)', '')
                        target_sheet.merge_cells(cell2)
                    target_xlsx.save(tagert_path)
                except Exception as e:
                    self.quit = True
                    QMessageBox.critical(self, "excel转换错误", f"{str(e)}")
                    break
            else:
                break
        if not self.quit:
            self.log.setText(f"导出目录:{tagert_path}")