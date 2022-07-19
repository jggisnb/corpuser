from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


def copy_xls(src_file, tag_file, sheet_name,mIndex):
    # src_file是源xlsx文件，tag_file是目标xlsx文件，sheet_name是目标xlsx里的新sheet名称
    src_wb = load_workbook(src_file)
    targ_wb = load_workbook(tag_file)
    src_ws = src_wb.worksheets[mIndex]
    targ_ws = targ_wb.create_sheet(sheet_name)
    max_row = src_ws.max_row  # 最大行数
    max_column = src_ws.max_column  # 最大列数
    w, h = 0, 0
    #此处有坑当你获得一个列宽为13的时候实际上是这个列和前面单元格一样的宽度，并不是他真的是13
    for i in range(1, max_column + 1):
        column_letter = get_column_letter(i)
        rs = src_ws.column_dimensions[column_letter].width
        if (rs == 13):
            rs = w
        else:
            w = rs
        targ_ws.column_dimensions[column_letter].width = rs
    #复制行高，没有列宽的坑
    for i in range(1, max_row + 1):
        rs = src_ws.row_dimensions[i].height
        if rs != None:
            targ_ws.row_dimensions[i].height = rs
    #复制每个单元格
    for column in range(1, max_column + 1):
        for row in range(1, max_row + 1):
            column_n = get_column_letter(column)
            i = '%s%d' % (column_n, row)  # 单元格编号
            try:
                #复制
                targ_ws[i].value = copy(src_ws[i].value)
                targ_ws[i].font = copy(src_ws[i].font)
                targ_ws[i].border = copy(src_ws[i].border)
                targ_ws[i].fill = copy(src_ws[i].fill)
                targ_ws[i].number_format = copy(src_ws[i].number_format)
                targ_ws[i].protection = copy(src_ws[i].protection)
                targ_ws[i].alignment = copy(src_ws[i].alignment)
            except Exception as e :
                print(e)

    # wm = list(src_ws.merged_cells)  # 开始处理合并单元格
    # for i in range(0, len(wm)):
    #     cell2 = str(wm[i]).replace('(<MergedCellRange ', '').replace('>,)', '')
    #     targ_ws.merge_cells(cell2)
    targ_wb.save(tag_file)
    targ_wb.close()
    src_wb.close()

def replace_xls(src_file, tag_file, sheet_name):
    #        src_file是源xlsx文件，tag_file是目标xlsx文件，sheet_name是目标xlsx里的新sheet名称

    print("Start sheet %s copy from %s to %s" % (sheet_name, src_file, tag_file))
    wb = load_workbook(src_file)
    wb2 = load_workbook(tag_file)

    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    ws2 = wb2.create_sheet(sheet_name)

    max_row = ws.max_row  # 最大行数
    max_column = ws.max_column  # 最大列数

    wm = list(ws.merged_cells)  # 开始处理合并单元格
    if len(wm) > 0:
        for i in range(0, len(wm)):
            cell2 = str(wm[i]).replace('(<MergeCell ', '').replace('>,)', '')
            print("MergeCell : %s" % cell2)
            ws2.merge_cells(cell2)

    for m in range(1, max_row + 1):
        ws2.row_dimensions[m].height = ws.row_dimensions[m].height
        for n in range(1, 1 + max_column):
            if n < 27:
                c = chr(n + 64).upper()  # ASCII字符,chr(65)='A'
            else:
                if n < 677:
                    c = chr(divmod(n, 26)[0] + 64) + chr(divmod(n, 26)[1] + 64)
                else:
                    c = chr(divmod(n, 676)[0] + 64) + chr(divmod(divmod(n, 676)[1], 26)[0] + 64) + chr(
                        divmod(divmod(n, 676)[1], 26)[1] + 64)
            i = '%s%d' % (c, m)  # 单元格编号
            if m == 1:
                #				 print("Modify column %s width from %d to %d" % (n, ws2.column_dimensions[c].width ,ws.column_dimensions[c].width))
                ws2.column_dimensions[c].width = ws.column_dimensions[c].width
            try:
                getattr(ws.cell(row=m, column=c), "value")
                cell1 = ws[i]  # 获取data单元格数据
                ws2[i].value = cell1.value  # 赋值到ws2单元格
                if cell1.has_style:  # 拷贝格式
                    ws2[i].font = copy(cell1.font)
                    ws2[i].border = copy(cell1.border)
                    ws2[i].fill = copy(cell1.fill)
                    ws2[i].number_format = copy(cell1.number_format)
                    ws2[i].protection = copy(cell1.protection)
                    ws2[i].alignment = copy(cell1.alignment)
            except AttributeError as e:
                print("cell(%s) is %s" % (i, e))
                continue

    wb2.save(tag_file)

    wb2.close()
    wb.close()